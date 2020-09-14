#  Copyright (c) 2019. Partners HealthCare and other members of
#  Forome Association
#
#  Developed by Sergey Trifonov based on contributions by Joel Krier,
#  Michael Bouzinier, Shamil Sunyaev and other members of Division of
#  Genetics, Brigham and Women's Hospital
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#        http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

import argparse
import json
import logging
import time
import os
from copy import copy

import openpyxl
from jsonpath_rw import parse


def cell_value(ws, row, column):
    v = ws.cell(row=row, column=column).value
    if (not v):
        return v
    return v.strip()


def read_mappings(fpath, verbose_mode):
    if (not os.path.isfile(fpath)):
        raise Exception("No Mapping file: %s" % fpath)
    if (not os.access(fpath, os.R_OK)):
        raise Exception("No read access to: %s" % fpath)

    if verbose_mode:
        logging.info("Reading: %s" % fpath)
    wb = openpyxl.load_workbook(fpath, read_only=False)

    mapping = _read_key_mapping(wb["key"], fpath)
    check_tags_mapping = _read_check_tags_mapping(wb["Check_Tags"],
        verbose_mode)
    wb.close()
    return mapping, check_tags_mapping

def _read_key_mapping(ws, fpath):
    if cell_value(ws, 1, 1) != "Column":
        raise Exception('First column must be called "Column". '
                        'Worksheet "key" of file %s' % fpath)
    key_column = 1
    map_column = getColumnByName(ws, "Mapping")
    if not map_column:
        raise Exception(
            "Column 'Mapping' is not found in Worksheet key of file %s"
            % fpath)
    def_column = getColumnByName(ws, "Definition")

    mapping = []
    for r in range(2, ws.max_row):
        cell = ws.cell(r, key_column)
        key = cell.value
        if not key:
            continue

        key = key.strip()
        style = dict()
        style["fill"] = copy(cell.fill)
        style["font"] = copy(cell.font)
        style["alignment"] = copy(cell.alignment)
        style["number_format"] = copy(cell.number_format)
        style["border"] = copy(cell.border)
        value = cell_value(ws, r, map_column)
        if value:
            def_value = cell_value(ws, r, def_column) if def_column else None
            mapping.append((len(mapping) + 1, key, value, style, def_value))
    return mapping

def _read_check_tags_mapping(ws, verbose_mode):
    key_column = 1
    res = {}
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, key_column)
        key = cell.value
        if verbose_mode:
            logging.info("tags: Reading key: %s" % key)
        if not key:
            continue

        key = key.strip()
        style = dict()
        style["fill"] = copy(cell.fill)
        style["font"] = copy(cell.font)
        style["alignment"] = copy(cell.alignment)
        style["number_format"] = copy(cell.number_format)
        if key:
            res[key] = style

    return res


def getColumnByName(ws, name):
    for c in range(1, ws.max_column):
        if cell_value(ws, 1, c) == name:
            return c
    return None

def build_value_jsonpath(array, key):
    # It's working, but very slowly

    jsonpath_expr = parse('$.."%s"' % key)
    match = jsonpath_expr.find(array)
    if match and match[0].value:
        if isinstance(match[0].value, list):
            return ','.join([str(item) for item in match[0].value])
        else:
            return match[0].value
    return None

def build_value(array, key):
    if array.get(key):
        return array[key]
    for x in array.values():
        value = None
        if isinstance(x, dict):
            value = build_value(x, key)
        elif isinstance(x, list):
            for element in x:
                if isinstance(element, dict):
                    value = build_value(element, key)
                    if value:
                        break
        if value:
            if isinstance(value, list):
                value = ','.join([str(item) for item in value])
            return value
    return None

def _setStyle(cell, style):
    if style:
        for s in style:
            setattr(cell, s, style[s])

class ExcelExport:
    def __init__(self, template_file, tags_info = None,
            source_versions = None, verbose_mode = False):
        self.mapping, self.check_tags_mapping = read_mappings(
            template_file, verbose_mode)
        self.workbook = None
        self.column_widths = {}
        self.tags_info = None
        self.check_group_tab = None
        self.add_tags_cfg(tags_info)
        if verbose_mode:
            for column in range(len(self.mapping)):
                logging.info("Column %s: %s"
                    % (column, self.mapping[column]))
        self.workbook = openpyxl.Workbook()
        self._createVariantSheet()
        self._createVersionSheet(source_versions)
        self._createKeySheet()

    def _createVariantSheet(self, title=None):
        ws = self.workbook.active
        ws.title = title if title else "Variants"
        for column, key, value, style, _ in self.mapping:
            if not value:
                continue
            cell = ws.cell(row = 1, column = column, value = key)
            self.column_widths[cell.column] = len(key)
            _setStyle(cell, style)

        for idx, title in enumerate(
                ["check tags", "tags", "tags with values", "notes"]):
            cell = ws.cell(row = 1, column = len(self.mapping) + 1 + idx,
                value = title)
            self.column_widths[cell.column] = len(cell.value)
        ws.freeze_panes = 'D2'

    def _createVersionSheet(self, source_versions):
        ws = self.workbook.create_sheet("version")
        if source_versions:
            for idx, pair in enumerate(source_versions):
                ws.cell(row=idx + 1, column = 1, value = pair[0])
                ws.cell(row=idx + 1, column = 2, value = pair[1])

    def _createKeySheet(self):
        ws = self.workbook.create_sheet("key")
        for idx, title in enumerate(["Column", "Definition", "Mapping"]):
            ws.cell(row = 1, column = idx + 1, value = title)
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(idx + 1)].width = 50
        ws.freeze_panes = 'A2'
        for row, key, value, style, def_value in self.mapping:
            if not value:
                continue
            cell = ws.cell(row = row + 1, column = 1, value = value)
            ws.cell(row = row + 1, column = 2, value = def_value)
            ws.cell(row = row + 1, column = 3, value = key)
            _setStyle(cell, style)

    def add_tags_cfg(self, data):
        if data is None:
            return
        self.tags_info = data
        self.check_group_tab = [0] * (len(self.tags_info['check-tags'])
            + len(self.tags_info['op-tags']) + 2)

    def reg_check_group(self, tags):
        if self.check_group_tab is None:
            return None, None
        group_idx = None
        group_name = None
        if tags is None:
            tags = dict()
        for check_idx, check_tag in enumerate(self.tags_info['check-tags']):
            if tags.get(check_tag):
                if group_idx is None:
                    group_idx, group_name = check_idx, check_tag
                else:
                    group_idx = len(self.check_group_tab) - 2
                    group_name = "_mix"
                    break
        for op_idx, op_tag in enumerate(self.tags_info['op-tags']):
            if op_tag == "_note":
                continue
            if op_tag in tags:
                if group_idx is None:
                    group_idx = len(self.tags_info['check-tags']) + op_idx
                    group_name = op_tag
                else:
                    group_idx = len(self.check_group_tab) - 2
                    group_name = "_mix"
                    break
        if group_idx is None:
            group_idx = len(self.check_group_tab) - 1
        self.check_group_tab[group_idx] += 1
        return group_name, 1 + sum(self.check_group_tab[:group_idx + 1])

    def add_variant(self, data, tags = None):
        ws = self.workbook.active
        tag_group_name, new_row = self.reg_check_group(tags)
        if new_row is None:
            new_row = ws.max_row + 1
        ws.insert_rows(new_row)
        for column, _, key, style, _ in self.mapping:
            if not key:
                continue
            value = self.__to_excel(build_value(data, key))
            cell = ws.cell(row = new_row, column = column, value = value)
            if isinstance(value, str):
                self.column_widths[cell.column] = max(
                    self.column_widths[cell.column], len(value))
            _setStyle(cell, style)
        if tags is not None and self.tags_info is not None:
            self.__add_tags_to_excel(tags, new_row, tag_group_name)

    def __add_tags_to_excel(self, tags, row, tag_group_name):
        ws = self.workbook.active
        check_tags, op_tags, tags_with_value = [], [], []
        for tg in self.tags_info['check-tags']:
            if tags.get(tg) is True:
                check_tags.append(tg)
        for tg in self.tags_info['op-tags']:
            if tg == "_note":
                continue
            val = tags.get(tg)
            if val is None:
                continue
            op_tags.append(tg)
            val = str(val).replace('\n', ' ').strip()
            if val:
                tags_with_value.append(tg + ': ' + val)
        note_value = tags.get("_note")
        if not note_value:
            note_value = ""
        for idx, value in enumerate([
                ', '.join(check_tags),
                ', '.join(op_tags),
                ', '.join(tags_with_value), note_value]):
            cell = ws.cell(row = row, column = len(self.mapping) + 1 + idx,
                value = value)
            self.column_widths[cell.column] = max(
                self.column_widths[cell.column], len(value))

        if tag_group_name:
            if tag_group_name in self.check_tags_mapping:
                style = self.check_tags_mapping[tag_group_name]
            elif tag_group_name == "_mix":
                style = self.check_tags_mapping["Multiple Tags"]
            else:
                style = self.check_tags_mapping["Custom"]
        else:
            style = None
        for idx in range(len(self.mapping) + 1):
            _setStyle(ws.cell(row = row, column = idx + 1), style)

    def _decor_one_line(self, ws, new_row, style = None):
        ws.insert_rows(new_row)
        for idx in range(len(self.mapping) + 1):
            _setStyle(ws.cell(row = new_row, column = idx + 1), style)

    def _decor_lines(self, ws):
        if self.check_group_tab is None:
            return
        if (self.check_group_tab[-1] > 0):
            cnt_before = sum(self.check_group_tab[:-1])
            if cnt_before > 0:
                self._decor_one_line(ws, cnt_before + 2)
        for idx in range(len(self.check_group_tab) - 2, -1, -1):
            if self.check_group_tab[idx] == 0:
                continue
            cnt_before = sum(self.check_group_tab[:idx])
            if idx >= len(self.tags_info['check-tags']):
                group_name = "_mix"
            else:
                group_name = self.tags_info['check-tags'][idx]
            self._decor_one_line(ws, cnt_before + 2,
                self.check_tags_mapping.get(group_name))

    def save(self, file):
        ws = self.workbook.active
        self._decor_lines(ws)
        for column, width in self.column_widths.items():
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(column)].width = min(
                    12, width + 2)
        max_column = openpyxl.utils.get_column_letter(ws.max_column)
        ws.auto_filter.ref = 'A1:' + max_column + str(len(ws['A']))
        self.workbook.save(filename=file)

    def __to_excel(self, value):
        if isinstance(value, str) and value.startswith("http"):
            return '=HYPERLINK("%s","%s")' % (value, value)
        if isinstance(value, dict):
            if "link" in value:
                return ('=HYPERLINK("%s","%s")'
                    % (value["link"], value.get("title", "")))
            return " ".join(sorted("%s=%s" % (key, val)
                for key, val in value.items()))
        return value


if __name__ == '__main__':
    import Enum

    class LoadMode(Enum):
        RECORD = "@RECORD"
        TAGS_CFG = "@TAGS_CFG"
        TAGS = "@TAGS"

    def processing(args):
        start_time = time.time()
        print("parsing template", args.template, "...")
        export = ExcelExport(args.template, verbose_mode = args.verbose)
        print("export variants from", args.input, "...")
        with open(args.input) as json_file:
            mode = LoadMode.RECORD
            record = None
            for idx, line in enumerate(json_file):
                if line.startswith("@"):
                    mode = LoadMode[line.strip()[1:]]
                else:
                    data = json.loads(line)
                    if mode == LoadMode.RECORD:
                        record = data
                    elif mode == LoadMode.TAGS_CFG:
                        export.add_tags_cfg(data)
                    elif mode == LoadMode.TAGS:
                        if record is not None:
                            export.add_variant(record, data)
                        record = None

                if args.limit and idx >= args.limit:
                    break
                if args.verbose and idx > 0 and idx % 100 == 0:
                    print("export lines:", idx)

            print("total export line:", idx)

        print("save", args.output)
        export.save(args.output)
        print("complete (execution time:",
            " %.03f s)" % (time.time() - start_time))

    parser = argparse.ArgumentParser()
    parser.add_argument("-t", "--template",
        help = "template file", required=True)
    parser.add_argument("-i", "--input",
        help = "input file with json lines", required=True)
    parser.add_argument("-o", "--output",
        help = "result file name", required=True)
    parser.add_argument("-l", "--limit",
        help = "maximum number of rows to export", type=int)
    parser.add_argument("-v", "--verbose",
        help = "increase output verbosity", action="store_true")
    args = parser.parse_args()
    processing(args)
